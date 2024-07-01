import os, sys
import PIL
from PIL import Image
import numpy as np
import numba
from numba import jit, prange, set_num_threads, threading_layer
import enum
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

counter = 1
green_color = PatternFill(patternType='solid', fgColor='35FC03')
red_color = PatternFill(patternType='solid', fgColor='FC2C03')

numba.config.THREADING_LAYER = 'threadsafe'
@jit(nopython=True)
def dither_floyd_steinberg_dev(image_array, w, h):
    copy_array = np.copy(image_array)
    w1 = 7 / 16
    w2 = 3 / 16
    w3 = 5 / 16
    w4 = 1 / 16
    for y in range(h):
        for x in range(w):
            old = copy_array[y, x]
            if old < 128:
                new = 0
            else:
                new = 255
            copy_array[y, x] = new
            error = old - new
            if y + 1 < h:
                copy_array[y + 1, x] += error * w1
            if (y + 1 < h) and (x + 1 < w):
                copy_array[y + 1, x + 1] += error * w4
            if x + 1 < w:
                copy_array[y, x + 1] += error * w3
            if (y - 1 >= 0) and (x + 1 < w):
                copy_array[y - 1, x + 1] += error * w2
    return copy_array
@jit(nopython=True)
def dither_atkinson(image_array, w, h):
    copy_array = np.copy(image_array)
    for y in range(h):
        for x in range(w):
            old = copy_array[y, x]
            if old < 128:
                new = 0
            else:
                new = 255
            copy_array[y, x] = new
            err = (old - new) >> 3  # divide by 8
            if x + 1 < w:
                copy_array[y, x + 1] += err
            if x + 2 < w:
                copy_array[y, x + 2] += err
            if (x - 1 < w) and (y + 1 < h):
                copy_array[y + 1, x - 1] += err
            if y + 1 < h:
                copy_array[y + 1, x] += err
            if (x + 1 < w) and (y + 1 < h):
                copy_array[y + 1, x + 1] += err
            if y + 2 < h:
                copy_array[y + 2, x] += err
    return copy_array


@jit(nopython=True)
def dither_jarvis_judice_ninke(image_array, w, h):
    copy_array = np.copy(image_array)
    w7 = 7 / 48
    w5 = 5 / 48
    w3 = 3 / 48
    w1 = 1 / 48
    for y in range(h):
        for x in range(w):
            old = copy_array[y, x]
            if old < 128:
                new = 0
            else:
                new = 255
            copy_array[y, x] = new
            err = (old - new)
            if x + 1 < w:
                copy_array[y, x + 1] += err * w7
            if x + 2 < w:
                copy_array[y, x + 2] += err * w5
            if (x - 2 >= 0) and (y + 1 < h):
                copy_array[y + 1, x - 2] += err * w3
            if (x - 1 >= 0) and (y + 1 < h):
                copy_array[y + 1, x - 1] += err * w5
            if y + 1 < h:
                copy_array[y + 1, x] += err * w7
            if (x + 1 < w) and (y + 1 < h):
                copy_array[y + 1, x + 1] += err * w5
            if (x + 2 < w) and (y + 1 < h):
                copy_array[y + 1, x + 2] += err * w3
            if (x - 2 >= 0) and (y + 2 < h):
                copy_array[y + 2, x - 2] += err * w1
            if (x - 1 >= 0) and (y + 2 < h):
                copy_array[y + 2, x - 1] += err * w3
            if y + 2 < h:
                copy_array[y + 2, x] += err * w5
            if (x + 1 < w) and (y + 2 < h):
                copy_array[y + 2, x + 1] += err * w3
            if (x + 2 < w) and (y + 2 < h):
                copy_array[y + 2, x + 2] += err * w1
    return copy_array
@jit(nopython=True)
def dither_stucki(image_array, w, h):
    copy_array = np.copy(image_array)
    w8 = 8 / 42.0
    w7 = 7 / 42.0
    w5 = 5 / 42.0
    w4 = 4 / 42.0
    w2 = 2 / 42.0
    w1 = 1 / 42.0
    for y in range(h):
        for x in range(w):
            old = copy_array[y, x]
            if old < 128:
                new = 0
            else:
                new = 255
            copy_array[y, x] = new
            err = (old - new)
            if x + 1 < w:
                copy_array[y, x + 1] += err * w8
            if x + 2 < w:
                copy_array[y, x + 2] += err * w4
            if (x - 2 >= 0) and (y + 1 < h):
                copy_array[y + 1, x - 2] += err * w2
            if (x - 1 >= 0) and (y + 1 < h):
                copy_array[y + 1, x - 1] += err * w4
            if y + 1 < h:
                copy_array[y + 1, x] += err * w8
            if (x + 1 < w) and (y + 1 < h):
                copy_array[y + 1, x + 1] += err * w4
            if (x + 2 < w) and (y + 1 < h):
                copy_array[y + 1, x + 2] += err * w2
            if (x - 2 >= 0) and (y + 2 < h):
                copy_array[y + 2, x - 2] += err * w1
            if (x - 1 >= 0) and (y + 2 < h):
                copy_array[y + 2, x - 1] += err * w2
            if y + 2 < h:
                copy_array[y + 2, x] += err * w4
            if (x + 1 < w) and (y + 2 < h):
                copy_array[y + 2, x + 1] += err * w2
            if (x + 2 < w) and (y + 2 < h):
                copy_array[y + 2, x + 2] += err * w1
    return copy_array

@jit(nopython=True)
def dither_bayer_2x2(image_array, w, h):
    copy_array = np.copy(image_array)
    dots = [2, 3, 4, 1]
    for k in range(len(dots)):
        dots[k] *= 64 - 1
    for y in prange(h):
        for x in range(w):
            if copy_array[y, x] >= dots[y % 2 * 2 + x % 2]:
                new = 255
            else:
                new = 0
            copy_array[y, x] = new
    return copy_array
def dither_bayer_8x8(image_array, w, h):
    copy_array = np.copy(image_array)
    dots = [1, 33, 9, 41, 3, 35, 11, 43, 49, 17, 57, 25, 51, 19, 59, 27, 13, 45, 5,
            37, 15, 47, 7, 39, 61, 29, 53, 21, 63, 31, 55, 23, 4, 36, 12, 44, 2, 34, 10, 42, 52,
            20, 60, 28, 50, 18, 58, 26, 16, 48, 8, 40, 14, 46, 6, 38, 64, 32, 56, 24, 62, 30,
            54, 22]
    for k in range(len(dots)):
        dots[k] *= 4 - 1
    for y in range(h):
        for x in range(w):
            if copy_array[y, x] >= dots[y % 8 * 8 + x % 8]:
                new = 255
            else:
                new = 0
            copy_array[y, x] = new
    return copy_array

@jit(nopython=True)
def dither_cluster_4x4(image_array, w, h):
    copy_array = np.copy(image_array)
    dots = [0.7500, 0.3750, 0.6250, 0.2500, 0.0625, 1.0000, 0.8750, 0.4375,
            0.5000, 0.8125, 0.9375, 0.1250, 0.1875, 0.5625, 0.3125, 0.6875]
    for k in range(len(dots)):
        dots[k] *= 255
    for y in range(h):
        for x in range(w):
            if copy_array[y, x] >= dots[y % 4 * 4 + x % 4]:
                new = 255
            else:
                new = 0
            copy_array[y, x] = new
    return copy_array


def recover_images_with_extension(folder_path, extension):
    image_files = []
    for file_name in os.listdir(folder_path):
        if file_name.endswith(f".{extension}") and os.path.isfile(os.path.join(folder_path, file_name)):
            image_files.append(os.path.join(folder_path, file_name))
    return image_files
def main(dir, method, output_dir, output_format, sheet):
    global counter
    sheet.cell(row=1, column=1).value = "image name"
    sheet.cell(row=1, column=2).value = "lossy compression method"
    sheet.cell(row=1, column=3).value = "output format"
    sheet.cell(row=1, column=4).value = "size (in bytes)"
    sheet.cell(row=1, column=5).value = "%"

    images = recover_images_with_extension(dir, "jpg")

    for image in images:
        im = Image.open(image)
        size_input_image_jpeg = os.stat(image).st_size
        sheet.cell(row=counter + 1, column=4).value = size_input_image_jpeg

        if method == "grayscale":
            # (8-bit pixels, black and white) grayscale
            # all values larger than 127 are set to 255 (white), all other values to 0 (black).
            im_L = im.convert("L", dither=Image.NONE)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "grayscale"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "grayscale"
            if output_format == "PNG":
                im_L.save(image_output_path + ".png" , "PNG")
                size_output_image_L = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_L
            elif output_format == "JPEG":
                im_L.save(image_output_path + ".jpg", "JPEG")
                size_output_image_L = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_L
            elif output_format == "WebP":
                im_L.save(image_output_path + ".WebP", "WebP")
                size_output_image_L = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_L
            else :
                 print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_L / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_L / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_L / size_input_image_jpeg) * 100

        elif method == "one_bit":
            im_1 = im.convert("1", dither=Image.NONE) # (1-bit pixels, black and white, stored with one pixel per byte)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "one_bit"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "one_bit"
            size_output_image_1 = 0
            if output_format == "PNG":
                im_1.save(image_output_path + ".png", "PNG")
                size_output_image_1 = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_1
            elif output_format == "JPEG":
                im_1.save(image_output_path + ".jpg", "JPEG")
                size_output_image_1 = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_1
            elif output_format == "WebP":
                im_1.save(image_output_path + ".WebP", "WebP")
                size_output_image_1 = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_1
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_1 / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_1 / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_1 / size_input_image_jpeg) * 100

        elif method == "floyd_steinberg_pil":
            im_F = im.convert("1", dither=Image.FLOYDSTEINBERG)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "floyd_steinberg_pil"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "floyd_steinberg_pil"
            size_output_image_F = 0

            if output_format == "PNG":
                im_F.save(image_output_path + ".png", "PNG")
                size_output_image_F = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_F
            elif output_format == "JPEG":
                im_F.save(image_output_path + ".jpg", "JPEG")
                size_output_image_F = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_F
            elif output_format == "WebP":
                im_F.save(image_output_path + ".WebP", "WebP")
                size_output_image_F = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_F
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_F / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_F / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_F / size_input_image_jpeg) * 100

        elif method == "floyd_steinberg_dev":
            im_f = im.convert("L", dither=Image.NONE)
            w, h = im_f.size
            image_array = np.asarray(im_f)
            image_array = dither_floyd_steinberg_dev(image_array, w, h)
            im_f = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "floyd_steinberg_dev"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "floyd_steinberg_dev"
            size_output_image_f = 0
            if output_format == "PNG":
                im_f.save(image_output_path + ".png", "PNG")
                size_output_image_f = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_f
            elif output_format == "JPEG":
                im_f.save(image_output_path + ".jpg", "JPEG")
                size_output_image_f = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_f
            elif output_format == "WebP":
                im_f.save(image_output_path + ".WebP", "WebP")
                size_output_image_f = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_f
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_f / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_f / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_f / size_input_image_jpeg) * 100

        elif method == "atkinson":
            im_a = im.convert("L", dither=Image.NONE)
            w, h = im_a.size
            image_array = np.asarray(im_a)
            image_array = dither_atkinson(image_array, w, h)
            im_a = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "atkinson"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "atkinson"
            size_output_image_a = 0
            if output_format == "PNG":
                im_a.save(image_output_path + ".png", "PNG")
                size_output_image_a = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_a
            elif output_format == "JPEG":
                im_a.save(image_output_path + ".jpg", "JPEG")
                size_output_image_a = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_a
            elif output_format == "WebP":
                im_a.save(image_output_path + ".WebP", "WebP")
                size_output_image_a = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_a
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_a / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_a / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_a / size_input_image_jpeg) * 100

        elif method == "jarvis_judice_ninke":
            im_j = im.convert("L", dither=Image.NONE)
            w, h = im_j.size
            image_array = np.asarray(im_j)
            image_array = dither_jarvis_judice_ninke(image_array, w, h)
            im_j = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "jarvis_judice_ninke"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "jarvis_judice_ninke"
            size_output_image_j = 0
            if output_format == "PNG":
                im_j.save(image_output_path + ".png", "PNG")
                size_output_image_j = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_j
            elif output_format == "JPEG":
                im_j.save(image_output_path + ".jpg", "JPEG")
                size_output_image_j = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_j
            elif output_format == "WebP":
                im_j.save(image_output_path + ".WebP", "WebP")
                size_output_image_j = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_j
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_j / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_j / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_j / size_input_image_jpeg) * 100

        elif method == "stucki":
            im_s = im.convert("L", dither=Image.NONE)
            w, h = im_s.size
            image_array = np.asarray(im_s)
            image_array = dither_stucki(image_array, w, h)
            im_s = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "stucki"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "stucki"
            size_output_image_s = 0
            if output_format == "PNG":
                im_s.save(image_output_path + ".png", "PNG")
                size_output_image_s = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_s
            elif output_format == "JPEG":
                im_s.save(image_output_path + ".jpg", "JPEG")
                size_output_image_s = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_s
            elif output_format == "WebP":
                im_s.save(image_output_path + ".WebP", "WebP")
                size_output_image_s = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_s
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_s / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_s / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_s / size_input_image_jpeg) * 100

        elif method == "bayer_2x2":
            im_b2 = im.convert("L", dither=Image.NONE)
            w, h = im_b2.size
            image_array = np.asarray(im_b2)
            image_array = dither_bayer_2x2(image_array, w, h)
            im_b2 = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "bayer_2x2"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "bayer_2x2"
            size_output_image_b2 = 0
            if output_format == "PNG":
                im_b2.save(image_output_path + ".png", "PNG")
                size_output_image_b2 = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b2
            elif output_format == "JPEG":
                im_b2.save(image_output_path + ".jpg", "JPEG")
                size_output_image_b2 = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b2
            elif output_format == "WebP":
                im_b2.save(image_output_path + ".WebP", "WebP")
                size_output_image_b2 = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b2
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_b2 / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_b2 / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_b2 / size_input_image_jpeg) * 100

        elif method == "bayer_8x8":
            im_b8 = im.convert("L", dither=Image.NONE)
            w, h = im_b8.size
            image_array = np.asarray(im_b8)
            image_array = dither_bayer_8x8(image_array, w, h)
            im_b8 = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "bayer_8x8"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "bayer_8x8"
            size_output_image_b8 = 0
            if output_format == "PNG":
                im_b8.save(image_output_path + ".png", "PNG")
                size_output_image_b8 = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b8
            elif output_format == "JPEG":
                im_b8.save(image_output_path + ".jpg", "JPEG")
                size_output_image_b8 = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b8
            elif output_format == "WebP":
                im_b8.save(image_output_path + ".WebP", "WebP")
                size_output_image_b8 = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_b8
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_b8 / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_b8 / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_b8 / size_input_image_jpeg) * 100

        elif method == "cluster_4x4":
            im_c = im.convert("L", dither=Image.NONE)
            w, h = im_c.size
            image_array = np.asarray(im_c)
            image_array = dither_cluster_4x4(image_array, w, h)
            im_c = Image.fromarray(image_array)
            image_name_only = os.path.basename(image)
            image_name_only = os.path.splitext(image_name_only)[0]
            image_output_path = output_dir + "/" + image_name_only + "_" + "cluster_4x4"
            sheet.cell(row=counter + 1, column=1).value = image_name_only
            sheet.cell(row=counter + 1, column=2).value = "bayer_8x8"
            size_output_image_c = 0
            if output_format == "PNG":
                im_c.save(image_output_path + ".png", "PNG")
                size_output_image_c = os.stat(image_output_path + ".png").st_size
                sheet.cell(row=counter + 2, column=3).value = "PNG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_c
            elif output_format == "JPEG":
                im_c.save(image_output_path + ".jpg", "JPEG")
                size_output_image_c = os.stat(image_output_path + ".jpg").st_size
                sheet.cell(row=counter + 2, column=3).value = "JPEG"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_c
            elif output_format == "WebP":
                im_c.save(image_output_path + ".WebP", "WebP")
                size_output_image_c = os.stat(image_output_path + ".WebP").st_size
                sheet.cell(row=counter + 2, column=3).value = "WebP"
                sheet.cell(row=counter + 2, column=4).value = size_output_image_c
            else:
                print("you didn't choose one of the available output format : PNG or JPEG or WebP")

            if ((size_output_image_c / size_input_image_jpeg) * 100 > 100):
                sheet.cell(row=counter + 2, column=5).fill = red_color
                sheet.cell(row=counter + 2, column=5).value = (size_output_image_c / size_input_image_jpeg) * 100
            else:
                sheet.cell(row=counter + 2, column=5).fill = green_color
                sheet.cell(row=counter + 2, column=5).value = 100 - (size_output_image_c / size_input_image_jpeg) * 100

        counter = counter + 5

if __name__ == "__main__":
    directory_path = sys.argv[1]
    name_excel_file = sys.argv[2]
    dithering_method = sys.argv[3]
    output_format = sys.argv[4]

    output_dir = directory_path + "_compressed"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = Workbook()
    if os.path.exists(output_dir + "/" + name_excel_file + ".xlsx"):
       wb.save(output_dir + "/" + name_excel_file + ".xlsx")
    sheet = wb.active
    if os.path.exists(directory_path):
        main(directory_path, dithering_method, output_dir, output_format, sheet)
    wb.save(output_dir + "/" + name_excel_file + ".xlsx")
